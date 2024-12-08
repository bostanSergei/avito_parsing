import openpyxl
from loguru import logger
from openpyxl.styles import Alignment

from datetime import date, datetime
import os

from settings import FILE


def get_path(flag: bool = True) -> str:
    if flag:
        folder = 'start_table'
    else:
        folder = 'table_after_parsing'

    if os.name == 'posix':
        path_to_file = f'{folder}/'
    else:
        path_to_file = f'{folder}\\'

    return path_to_file


def get_start_data_from_excel_table(name: str = 'sample') -> dict:
    """
    Данный метод позволит открыть стартовый excel документ для получения необходимых данных для последующего парсинга.
    # name - имя стартовой таблицы (по умолчанию 'sample.xlsx')

    UPD от 01.12.2024 - формат стартовой таблицы изменяется! Подробнее в файле state.txt в корне
    """
    result_data: dict = {
        'search_query': None,       # предполагается строка для последующей отправки запроса
        'categories': None,         # предполагается список с категориями, где первый элемент - верхний уровень
        'cities': None,             # предполагается словарь, где ключ - это номер ячейки, в которой находится город,
                                    # а значение - сам город (этот финт нужен для последующей записи полученных данных)
        'company_name': None,       # название компании, которое будет учитываться в поиске
    }

    if f'{name}.xlsx' not in os.listdir(f'start_table/' if os.name == 'posix' else f'start_table\\'):
        logger.error('ТАБЛИЦА НЕ НАЙДЕНА')
        raise FileNotFoundError

    workbook = openpyxl.load_workbook(f'{get_path()}{name}.xlsx')
    sheet = workbook.active

    logger.info(f'Открыта стартовая таблица с именем {name}.xlsx')

    # в теории может возникнуть ситуация, в которой название компании отсутствует. В этом случае мы можем пропустить E-колонку
    result_data['company_name'] = sheet['B2'].value.strip() if sheet['B2'].value is not None else None

    # изменили строки исходных данных
    # result_data['search_query'] = sheet['B3'].value.strip()
    result_data['search_query'] = sheet['B4'].value.strip()

    # result_data['categories'] = [value.strip() for value in sheet['B2'].value.split('\n')]
    result_data['categories'] = [value.strip() for value in sheet['B3'].value.split('\n')]

    # start_column, start_row = 'A', 8
    start_column, start_row = 'A', 9
    column_values: dict = {}
    uniq_cities: set = set()

    for row in range(start_row, sheet.max_row + 1):
        cur_value = sheet[f'{start_column}{row}'].value
        if cur_value is None:
            break
        else:
            final_string = cur_value
            if final_string not in uniq_cities:
                uniq_cities.add(final_string)
                column_values[f'{row}'] = final_string

    result_data['cities'] = column_values
    workbook.close()

    logger.info('Данные извлечены из таблицы и переданы на следующий уровень')
    return result_data


def save_excel_table(final_data: dict, company_name: str) -> str:
    """
    UPD переписываем сохранение таблицы от 01.12.2024
    Теперь final_data - это словарь, в котором ключи - это номера строк, а значения - это словари с результатом парсинга
    """
    logger.info('Открываем таблицу для сохранения данных по результатам поисковых запросов')
    workbook = openpyxl.load_workbook(filename=f'{get_path()}{FILE}.xlsx')
    sheet = workbook.active

    # ПРИМЕР ОТВЕТА - final_result - это вложенный словарь, соответствующий номеру строки
    # final_result = {
    #     'all_ads': None,
    #     'promoted_ads': None,
    #     'companies': {
    #         "LaserSkin | Косметологическое оборудование.": {
    #             "count": 2,
    #             "places": [
    #                 1,
    #                 2
    #             ],
    #             "pay_places": [
    #                 1,
    #                 2
    #             ]
    #         },
    #         "Best_Beauty_Shop": {
    #             "count": 2,
    #             "places": [
    #                 3,
    #                 31
    #             ],
    #             "pay_places": []
    #         },
    #     }
    # }

    company_name = company_name.strip()
    for row, current_dict in final_data.items():
        if current_dict is None:
            sheet[f'C{row}'] = -1
            sheet[f'D{row}'] = -1
            continue

        print(current_dict)
        sheet[f'C{row}'] = current_dict['all_ads']
        sheet[f'D{row}'] = current_dict['promoted_ads']

        if company_name in current_dict['companies']:
            positions = '\n'.join(map(str, current_dict['companies'][company_name]['places']))
            sheet[f'E{row}'] = positions
        else:
            sheet[f'E{row}'] = 0

        all_companies = []
        for key, value in current_dict['companies'].items():
            all_companies.append(
                f'{key} - {value["count"]} ({len(value["pay_places"])})'
            )
        if len(all_companies) > 0:
            sheet[f'F{row}'] = '\n'.join(all_companies)
            sheet[f'F{row}'].alignment = Alignment(wrap_text=True)

    curr_date = date.today().strftime('%d_%m_%Y')

    workbook.save(filename=f'{get_path(False)}parsing_{curr_date}.xlsx')
    workbook.close()

    logger.info(f'Таблица сохранена и закрыта с именем: parsing_{curr_date}.xlsx')
    return curr_date


def save_analytic_part(final_data: dict, search_query: list, curr_date: str) -> None:
    logger.info('Открываем таблицу для сохранения данных по результатам аналитики')
    workbook = openpyxl.load_workbook(filename=f'{get_path(False)}parsing_{curr_date}.xlsx')
    sheet = workbook.active

    for row, lst in final_data.items():
        if len(lst) == len(search_query):
            string = ''
            for i in range(len(search_query)):
                string += f'{search_query[i]} - {str(lst[i])}\n'

            sheet[f'B{row}'] = string.strip()

    curr_date = datetime.now().strftime('%d_%m_%Y_%H_%M')
    workbook.save(filename=f'{get_path(False)}parsing_{curr_date}.xlsx')
    workbook.close()

    logger.info(f'Таблица сохранена и закрыта с именем: parsing_{curr_date}.xlsx')


def get_analytic_data() -> dict:
    logger.info('Открываем таблицу для получения данных на этап аналитики')
    path_to_file = f'{get_path(flag=True)}{FILE}.xlsx'

    result_data: dict = {
        'search_query': None,       # предполагается список со строками для поискового запроса
        'category': None,           # предполагается строка с категорией для поиска
        'cities': None              # предполагается словарь, где ключ - это номер ячейки, в которой находится город,
                                    # а значение - сам город (этот финт нужен для последующей записи полученных данных)
                                    # к тому же в словаре будут отсутствовать города, не найденные на предыдущем шаге
    }

    workbook = openpyxl.load_workbook(f'{path_to_file}')
    sheet = workbook.active

    # result_data['search_query'] = [value.strip().lower() for value in sheet['B4'].value.split('\n')]
    # result_data['category'] = sheet['B5'].value.strip().lower()

    # изменили строки исходных данных
    result_data['search_query'] = [value.strip().lower() for value in sheet['B5'].value.split('\n')]
    result_data['category'] = sheet['B6'].value.strip().lower()

    # start_column, start_row = 'A', 8
    start_column, start_row = 'A', 9
    column_values: dict = {}

    for row in range(start_row, sheet.max_row + 1):
        cur_value = sheet[f'{start_column}{row}'].value
        if cur_value is None:
            break
        else:
            final_string = cur_value
            # if (curr_data := sheet[f'C{row}'].value) is not None and str(curr_data).isdigit():
            column_values[f'{row}'] = final_string.strip()

    result_data['cities'] = column_values
    workbook.close()

    logger.info('Данные сохранены - таблица закрыта!')
    return result_data
